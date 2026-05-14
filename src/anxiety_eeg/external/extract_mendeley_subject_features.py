"""中文说明

用途：
    从 Mendeley anxiety/control Excel 工作簿中提取与主模型部分兼容的受试者级
    theta/alpha/beta 频谱特征。
输入：
    `--workbook` 指向官方下载的 `EEG_data.xlsx`，默认查找
    `data/mendeley_anxiety_control/EEG_data.xlsx`。
输出：
    `subject_features.csv`、`dataset_info.json`、兼容性摘要和失败记录。
快速运行：
    `python -m anxiety_eeg.external.extract_mendeley_subject_features --workbook data/mendeley_anxiety_control/EEG_data.xlsx`
论文对应：
    第 5 章 Mendeley 外部兼容性验证。
注意事项：
    Mendeley 缺少 delta、gamma 与 beta/(delta+theta)，只能作为 partial compatibility。
"""

from __future__ import annotations

import argparse
import json
import math
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


from anxiety_eeg.features.common_dataset_eval import SubjectFeatureBundle, save_dataset_outputs  # noqa: E402


REPO_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_WORKBOOK = REPO_ROOT / "data" / "mendeley_anxiety_control" / "EEG_data.xlsx"
DEFAULT_OUTPUT_DIR = REPO_ROOT / "features" / "external" / "mendeley"
DEFAULT_DATASET_NAME = "mendeley_anxiety_control"
DEFAULT_SCORE_NAME = "group_binary"
EPS = 1e-8
ABS_CONDITION_COLUMNS = {
    "observ": 6,
    "imagin": 7,
    "execut": 8,
    "together": 9,
}
SOURCE_BAND_TO_TARGET = {
    "theta2": "theta",
    "alpha1": "alpha",
    "alpha2": "alpha",
    "beta1": "beta",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract subject-level compatibility features from the Mendeley anxiety/control workbook."
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--dataset-name", type=str, default=DEFAULT_DATASET_NAME)
    parser.add_argument("--score-name", type=str, default=DEFAULT_SCORE_NAME)
    parser.add_argument(
        "--condition",
        choices=["observ", "imagin", "execut", "together", "mean_tasks"],
        default="together",
        help="Which workbook task column to summarize into one subject row.",
    )
    parser.add_argument("--min-n", type=int, default=4)
    return parser.parse_args()


def finite_float(value: object) -> float | None:
    if value is None:
        return None
    try:
        out = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(out):
        return None
    return out


def detect_final_subject_sheet(workbook) -> str:
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        normalized = [("" if value is None else str(value).strip().lower()) for value in header]
        if len(normalized) >= 21 and normalized[1:5] == ["code", "group", "area", "channel"] and "theta2" in normalized:
            return sheet_name
    raise RuntimeError("Could not detect the final Mendeley subject sheet with code/group/channel/theta2 columns.")


def detect_positive_band_sheet(workbook) -> str:
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        sample = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
        if not sample or len(sample) < 10:
            continue
        band = "" if sample[5] is None else str(sample[5]).strip().lower()
        condition_value = finite_float(sample[6])
        if band in SOURCE_BAND_TO_TARGET and condition_value is not None and condition_value >= 0.0:
            return sheet_name
    raise RuntimeError("Could not detect the positive-power Mendeley source sheet.")


def normalize_group(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip().lower()
    if text in {"control", "controls"}:
        return "control"
    if text in {"patient", "patients"}:
        return "patients"
    return None


def normalize_channel(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if "-" in text:
        text = text.split("-", 1)[0].strip()
    return text


def select_absolute_value(row: tuple, condition: str) -> float | None:
    if condition == "mean_tasks":
        values = []
        for column_index in ABS_CONDITION_COLUMNS.values():
            value = finite_float(row[column_index] if len(row) > column_index else None)
            if value is not None and value > 0.0:
                values.append(float(value))
        return None if not values else float(sum(values) / len(values))
    column_index = ABS_CONDITION_COLUMNS[condition]
    value = finite_float(row[column_index] if len(row) > column_index else None)
    if value is None or value <= 0.0:
        return None
    return float(value)


def load_final_subject_map(workbook, sheet_name: str) -> dict[str, str]:
    ws = workbook[sheet_name]
    out: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        subject = "" if row[1] is None else str(row[1]).strip()
        group = normalize_group(row[2])
        if not subject or group is None:
            continue
        out[subject] = group
    if not out:
        raise RuntimeError(f"No subjects found in final subject sheet: {sheet_name}")
    return out


def build_bundles(
    workbook,
    positive_sheet_name: str,
    final_subject_map: dict[str, str],
    dataset_name: str,
    score_name: str,
    condition: str,
) -> tuple[list[SubjectFeatureBundle], list[dict], dict]:
    ws = workbook[positive_sheet_name]
    per_subject_channel_band: dict[str, dict[str, dict[str, float]]] = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    failures: list[dict] = []
    raw_row_counter = 0
    skipped_nonfinal = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_row_counter += 1
        if not row or len(row) < 10:
            continue
        subject = "" if row[1] is None else str(row[1]).strip()
        if not subject:
            continue
        if subject not in final_subject_map:
            skipped_nonfinal += 1
            continue
        band = "" if row[5] is None else str(row[5]).strip().lower()
        target_band = SOURCE_BAND_TO_TARGET.get(band)
        if target_band is None:
            continue
        channel = normalize_channel(row[2])
        if channel is None:
            failures.append({"subject": subject, "reason": "invalid_channel_name", "raw_channel": row[2]})
            continue
        abs_value = select_absolute_value(row, condition)
        if abs_value is None:
            failures.append({"subject": subject, "reason": "missing_abs_power", "channel": channel, "band": band})
            continue
        per_subject_channel_band[subject][channel][target_band] += float(abs_value)

    bundles: list[SubjectFeatureBundle] = []
    label_counter = Counter()
    channel_counter = Counter()

    for subject, channel_map in sorted(per_subject_channel_band.items()):
        group = final_subject_map[subject]
        label_counter[group] += 1
        anxiety = 1.0 if group == "patients" else 0.0

        channel_features: dict[str, dict[str, float]] = {}
        for channel, band_map in sorted(channel_map.items()):
            theta = float(band_map.get("theta", 0.0))
            alpha = float(band_map.get("alpha", 0.0))
            beta = float(band_map.get("beta", 0.0))
            total = theta + alpha + beta
            if total <= EPS:
                continue
            channel_features[channel] = {
                "rel_theta": float(theta / total),
                "rel_alpha": float(alpha / total),
                "rel_beta": float(beta / total),
                "log_theta": float(math.log(max(theta, EPS))),
                "log_alpha": float(math.log(max(alpha, EPS))),
                "log_beta": float(math.log(max(beta, EPS))),
            }

        if not channel_features:
            failures.append({"subject": subject, "reason": "no_usable_channels_after_band_merge"})
            continue

        channel_names = sorted(channel_features)
        channel_counter.update(channel_names)
        bundles.append(
            SubjectFeatureBundle(
                dataset_name=dataset_name,
                score_name=score_name,
                subject=subject,
                split="external_binary_group",
                anxiety=float(anxiety),
                context=condition,
                channel_names=channel_names,
                channel_features=channel_features,
                source_count=1,
                metadata={
                    "group": group,
                    "source_sheet": positive_sheet_name,
                    "condition": condition,
                    "feature_mode": "overlap_only_abs_power",
                },
            )
        )

    summary = {
        "dataset_name": dataset_name,
        "score_name": score_name,
        "condition": condition,
        "positive_sheet": positive_sheet_name,
        "n_final_subjects": int(len(final_subject_map)),
        "n_subjects_with_features": int(len(bundles)),
        "label_counts": {key: int(value) for key, value in sorted(label_counter.items())},
        "raw_rows_scanned": int(raw_row_counter),
        "skipped_nonfinal_rows": int(skipped_nonfinal),
        "top_channels": channel_counter.most_common(12),
    }
    return bundles, failures, summary


def write_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> int:
    args = parse_args()
    workbook_path = args.workbook.resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Mendeley workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    final_sheet_name = detect_final_subject_sheet(workbook)
    positive_sheet_name = detect_positive_band_sheet(workbook)
    final_subject_map = load_final_subject_map(workbook, final_sheet_name)
    bundles, failures, summary = build_bundles(
        workbook=workbook,
        positive_sheet_name=positive_sheet_name,
        final_subject_map=final_subject_map,
        dataset_name=args.dataset_name,
        score_name=args.score_name,
        condition=args.condition,
    )

    output_dir = args.output_dir.resolve()
    config = {
        "workbook": str(workbook_path),
        "condition": args.condition,
        "final_subject_sheet": final_sheet_name,
        "positive_power_sheet": positive_sheet_name,
        "label_mode": "binary_group_from_workbook",
        "band_mapping": SOURCE_BAND_TO_TARGET,
        "missing_training_features": [
            "global_rel_delta",
            "global_rel_gamma",
            "global_beta_over_delta_theta",
        ],
    }
    payload = save_dataset_outputs(
        dataset_name=args.dataset_name,
        score_name=args.score_name,
        bundles=bundles,
        failures=failures,
        source_root=workbook_path.parent,
        output_dir=output_dir,
        config=config,
        min_n=int(args.min_n),
    )

    compatibility = {
        **summary,
        "subject_feature_csv": str((output_dir / "subject_features.csv").resolve()),
        "feature_names": sorted(
            key for key in payload["subject_rows"][0].keys() if key not in {"dataset", "score_name", "subject", "split", "context", "anxiety", "n_channels", "source_count"}
        )
        if payload["subject_rows"]
        else [],
        "dataset_info_path": str((output_dir / "dataset_info.json").resolve()),
        "notes": [
            "Only overlapping theta/alpha/beta information is extracted from the workbook.",
            "delta/gamma-dependent training inputs are intentionally left absent here and should be neutral-imputed during external load.",
            "Workbook labels are binary group labels, not subject-level continuous anxiety scores.",
        ],
    }
    write_json(output_dir / "compatibility_summary.json", compatibility)

    print(f"[Mendeley] workbook={workbook_path}")
    print(f"[Mendeley] final_subject_sheet={final_sheet_name}")
    print(f"[Mendeley] positive_power_sheet={positive_sheet_name}")
    print(f"[Mendeley] subjects_with_features={summary['n_subjects_with_features']} / {summary['n_final_subjects']}")
    print(f"[Mendeley] subject_features={output_dir / 'subject_features.csv'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
